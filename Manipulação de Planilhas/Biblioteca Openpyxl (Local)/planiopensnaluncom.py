from openpyxl import load_workbook

caminho = 'applan.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active

conteudo = planilha["B5"]
planilha.cell(row=7, column=4, value='learn')
cham = {}
presen = {}
presen['SIM'] = 0
presen['NÃO'] = 0
cham2 = {}
cham3 = {}
for c in range(5,36):
    if planilha.cell(row=c, column=2).value == None or len(planilha.cell(row=c, column=2).value) == 14:
        j = 0
    else:
        if planilha.cell(row=c, column=2).value not in cham.keys():
            cham[planilha.cell(row=c, column=2).value] = presen.copy()
            if planilha.cell(row=c, column=3).value[0] == "S":
                t = cham[planilha.cell(row=c, column=2).value].get('SIM')
                t += 1
                cham[planilha.cell(row=c, column=2).value].update({'SIM': t})
            elif planilha.cell(row=c, column=3).value[0] == "N":
                f = cham[planilha.cell(row=c, column=2).value].get('NÃO')
                f += 1
                cham[planilha.cell(row=c, column=2).value].update({'NÃO': f})
        else:
            if planilha.cell(row=c, column=3).value[0] == "S":
                t = cham[planilha.cell(row=c, column=2).value].get('SIM')
                t += 1
                cham[planilha.cell(row=c, column=2).value].update({'SIM': t})
            elif planilha.cell(row=c, column=3).value[0] == "N":
                f = cham[planilha.cell(row=c, column=2).value].get('NÃO')
                f += 1
                cham[planilha.cell(row=c, column=2).value].update({'NÃO': f})
for d in range(22,28):
    if planilha.cell(row=d, column=2).value not in cham2.keys():
        cham2[planilha.cell(row=d, column=2).value] = []
    if d == 27:
        for o in cham.keys():
            if o not in cham2.keys():
                p = cham[o].get('NÃO')
                p += 1
                cham[o].update({'NÃO': p})

for i in range(32,36):
    if planilha.cell(row=i, column=2).value not in cham3.keys():
        cham3[planilha.cell(row=i, column=2).value] = []
    if i == 35:
        for y in cham.keys():
            if y not in cham3.keys():
                l = cham[y].get('NÃO')
                l += 1
                cham[y].update({'NÃO': l})

for u in range(42,53):
    if planilha.cell(row=u, column=5).value in cham.keys():
        s = cham[planilha.cell(row=u, column=5).value].get('SIM')
        n = cham[planilha.cell(row=u, column=5).value].get('NÃO')

        planilha.cell(row=u, column=6).value = s
        planilha.cell(row=u, column=7).value = n
print(cham)
arquivo_excel.save("applan.xlsx")